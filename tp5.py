import re
import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

sys.stdout.reconfigure(encoding="utf-8")

# Ruta del archivo CSV (misma carpeta que el script)
ARCHIVO_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "export-2019-to-now-v4.csv")

# Expresiones regulares para validar cada campo del CSV.
# Si un campo no cumple su regex, el registro se descarta.
REGEX = {
    "ID":               re.compile(r"^\d+$"),
    "ID_Sesion":        re.compile(r"^[0-9A-Fa-f]{8}-[0-9A-Fa-f]{8}$"),
    "ID_Conexion":      re.compile(r"^[0-9a-f]{16}$"),
    "Usuario":          re.compile(r"^[A-Za-z0-9._\-]+$"),
    "IP":               re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$"),
    "Tipo_conexion":    re.compile(r"^[\w\-\.]+$"),
    "Fecha":            re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    "Hora":             re.compile(r"^\d{2}:\d{2}:\d{2}$"),
    "Numerico":         re.compile(r"^\d+$"),
    "MAC_AP":           re.compile(r"^([0-9A-Fa-f]{2}-){5}[0-9A-Fa-f]{2}(:\w+)?$"),
    "MAC_Cliente":      re.compile(r"^([0-9A-Fa-f]{2}-){5}[0-9A-Fa-f]{2}$"),
    "Razon":            re.compile(r"^[\w\-]+$"),
}

# Orden de validación: (índice en el CSV, nombre del campo, clave en REGEX)
CAMPOS = [
    (0,  "ID",              "ID"),
    (1,  "ID_Sesion",       "ID_Sesion"),
    (2,  "ID_Conexion",     "ID_Conexion"),
    (3,  "Usuario",         "Usuario"),
    (4,  "IP_NAS_AP",       "IP"),
    (5,  "Tipo_conexion",   "Tipo_conexion"),
    (6,  "Inicio_Dia",      "Fecha"),
    (7,  "Inicio_Hora",     "Hora"),
    (8,  "Fin_Dia",         "Fecha"),
    (9,  "Fin_Hora",        "Hora"),
    (10, "Session_Time",    "Numerico"),
    (11, "Input_Octects",   "Numerico"),
    (12, "Output_Octects",  "Numerico"),
    (13, "MAC_AP",          "MAC_AP"),
    (14, "MAC_Cliente",     "MAC_Cliente"),
    (15, "Razon",           "Razon"),
]


def validar_registro(fila):
    """Valida los 16 campos de un registro con expresiones regulares.
    Retorna (True, None) si es válido o (False, 'campo con error') si no."""
    for indice, nombre, clave in CAMPOS:
        valor = fila[indice].strip()
        if not valor or not REGEX[clave].match(valor):
            return False, nombre
    return True, None


def cargar_csv():
    """Lee el CSV, valida cada registro y separa válidos de inválidos."""
    validos = []
    invalidos = []

    with open(ARCHIVO_CSV, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        next(reader)  # Saltamos encabezado

        for num_linea, fila in enumerate(reader, start=2):
            if len(fila) < 16:
                invalidos.append((num_linea, "Campos insuficientes"))
                continue

            ok, campo_error = validar_registro(fila)
            if ok:
                validos.append({
                    "Usuario":      fila[3].strip(),
                    "Inicio_Dia":   fila[6].strip(),
                    "Session_Time": int(fila[10].strip()),
                })
            else:
                invalidos.append((num_linea, campo_error))

    return validos, invalidos


def filtrar_por_fecha(registros, fecha_inicio, fecha_fin):
    """Filtra registros cuya fecha de inicio esté en el rango dado.
    Funciona con comparación de strings porque YYYY-MM-DD mantiene orden."""
    return [r for r in registros if fecha_inicio <= r["Inicio_Dia"] <= fecha_fin]


def agrupar_por_usuario(registros):
    """Agrupa registros por usuario, sumando tiempo y contando sesiones."""
    usuarios = {}
    for r in registros:
        u = r["Usuario"]
        if u not in usuarios:
            usuarios[u] = {"tiempo": 0, "sesiones": 0}
        usuarios[u]["tiempo"] += r["Session_Time"]
        usuarios[u]["sesiones"] += 1
    return usuarios


def formatear_tiempo(segundos):
    """Convierte segundos a formato legible (ej: 2d 5h 30m 10s)."""
    d = segundos // 86400
    h = (segundos % 86400) // 3600
    m = (segundos % 3600) // 60
    s = segundos % 60
    partes = []
    if d > 0: partes.append(f"{d}d")
    if h > 0: partes.append(f"{h}h")
    if m > 0: partes.append(f"{m}m")
    if s > 0 or not partes: partes.append(f"{s}s")
    return " ".join(partes)


def pedir_fechas():
    """Pide al usuario un rango de fechas. Ofrece el rango de pandemia por defecto."""
    print("\nRango por defecto (pandemia COVID-19): 2020-03-20 a 2021-12-31")
    opcion = input("¿Usar rango por defecto? (S/n): ").strip().lower()

    if opcion in ("", "s", "si", "sí"):
        return "2020-03-20", "2021-12-31"

    regex_fecha = re.compile(r"^\d{4}-\d{2}-\d{2}$")

    while True:
        inicio = input("Fecha de INICIO (YYYY-MM-DD): ").strip()
        if regex_fecha.match(inicio):
            break
        print("Formato inválido.")

    while True:
        fin = input("Fecha de FIN (YYYY-MM-DD): ").strip()
        if regex_fecha.match(fin) and fin >= inicio:
            break
        print("Formato inválido o fecha anterior al inicio.")

    return inicio, fin


def mostrar_resultados(ordenados, fecha_inicio, fecha_fin):
    """Muestra la tabla de usuarios ordenados por tiempo descendente."""
    print(f"\n{'=' * 65}")
    print(f"  USUARIOS CONECTADOS: {fecha_inicio} a {fecha_fin}")
    print(f"  Total de usuarios: {len(ordenados)}")
    print(f"{'=' * 65}")

    if not ordenados:
        print("  No se encontraron usuarios en ese rango.")
        return

    print(f"\n  {'#':<5} {'Usuario':<28} {'Tiempo Total':<18} {'Sesiones':<10}")
    print(f"  {'-'*5} {'-'*28} {'-'*18} {'-'*10}")

    for i, (usuario, datos) in enumerate(ordenados, 1):
        print(f"  {i:<5} {usuario:<28} {formatear_tiempo(datos['tiempo']):<18} {datos['sesiones']:<10}")


def mostrar_descartados(invalidos):
    """Muestra los registros descartados por errores de validación."""
    print(f"\n{'=' * 65}")
    print(f"  REGISTROS DESCARTADOS: {len(invalidos)}")
    print(f"{'=' * 65}")

    if not invalidos:
        print("  Todos los registros son válidos.")
        return

    limite = min(len(invalidos), 30)
    print(f"\n  Mostrando {limite} de {len(invalidos)}:\n")
    print(f"  {'Línea':<10} {'Campo con error':<30}")
    print(f"  {'-'*10} {'-'*30}")
    for linea, error in invalidos[:limite]:
        print(f"  {linea:<10} {error:<30}")


def exportar_excel(ordenados, fecha_inicio, fecha_fin):
    """Exporta los resultados a un archivo Excel (.xlsx)."""
    if not OPENPYXL_DISPONIBLE:
        print("\n  El módulo 'openpyxl' no está instalado.")
        print("  Para instalarlo ejecute: pip install openpyxl")
        return

    nombre = f"usuarios_covid19_{fecha_inicio}_a_{fecha_fin}.xlsx"
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), nombre)

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios COVID-19"

    # Estilos
    fuente_titulo = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    fuente_encabezado = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    relleno_titulo = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    relleno_encabezado = PatternFill(start_color="048A81", end_color="048A81", fill_type="solid")
    centro = Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells("A1:D1")
    ws["A1"].value = f"Usuarios conectados ({fecha_inicio} a {fecha_fin})"
    ws["A1"].font = fuente_titulo
    ws["A1"].fill = relleno_titulo
    ws["A1"].alignment = centro

    # Encabezados
    for col, texto in enumerate(["#", "Usuario", "Tiempo Total", "Sesiones"], 1):
        celda = ws.cell(row=3, column=col, value=texto)
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = centro

    # Datos
    for i, (usuario, datos) in enumerate(ordenados, 1):
        ws.cell(row=i + 3, column=1, value=i)
        ws.cell(row=i + 3, column=2, value=usuario)
        ws.cell(row=i + 3, column=3, value=formatear_tiempo(datos["tiempo"]))
        ws.cell(row=i + 3, column=4, value=datos["sesiones"])

    # Ancho de columnas
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12

    wb.save(ruta)
    print(f"\n  Archivo exportado: {nombre}")


# =============================================================================
# PROGRAMA PRINCIPAL
# =============================================================================

if __name__ == "__main__":
    print("=" * 65)
    print("  TP5 - TRABAJO INTEGRADOR DE EXPRESIONES REGULARES")
    print("  Ejercicio 8: Usuarios conectados durante pandemia COVID-19")
    print("=" * 65)

    # Verificar que el CSV existe
    if not os.path.exists(ARCHIVO_CSV):
        print(f"\nError: No se encontró el archivo CSV.")
        sys.exit(1)

    # Cargar y validar
    print("\nCargando y validando registros...")
    validos, invalidos = cargar_csv()
    print(f"  Válidos:   {len(validos):,}")
    print(f"  Inválidos: {len(invalidos):,}")

    # Menú
    ultima_consulta = None
    ultima_fecha = None

    while True:
        print(f"\n{'-' * 40}")
        print("  1. Consultar por rango de fechas")
        print("  2. Ver registros descartados")
        print("  3. Exportar última consulta a Excel")
        print("  4. Salir")
        print(f"{'-' * 40}")

        opcion = input("  Opción: ").strip()

        if opcion == "1":
            fecha_inicio, fecha_fin = pedir_fechas()
            filtrados = filtrar_por_fecha(validos, fecha_inicio, fecha_fin)
            agrupados = agrupar_por_usuario(filtrados)
            # Ordenar por tiempo descendente
            ultima_consulta = sorted(agrupados.items(), key=lambda x: x[1]["tiempo"], reverse=True)
            ultima_fecha = (fecha_inicio, fecha_fin)
            mostrar_resultados(ultima_consulta, fecha_inicio, fecha_fin)

        elif opcion == "2":
            mostrar_descartados(invalidos)

        elif opcion == "3":
            if ultima_consulta is None:
                print("\n  Primero realice una consulta (opción 1).")
            else:
                exportar_excel(ultima_consulta, *ultima_fecha)

        elif opcion == "4":
            print("\n¡Hasta luego!")
            break
